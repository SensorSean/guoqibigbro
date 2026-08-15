"""
填表执行引擎 V2（重写版）

核心改进（相对于 guoqi-bigbro 的 V1 顺序填充）：
  - 行标识(key)匹配：按行标识（如姓名、编号）对齐源和目标的"行"，而非顺序填充
  - 多源文件优先级：后添加的源文件数据覆盖先添加的（同key同列取最后值）
  - 空模板自动收集：目标模板无数据行时，自动从所有源文件收集唯一行标识
  - 自动检测 key 列：未指定 key_column 时，自动扫描非空值最多的列
  - openpyxl 写回保留格式：保存结果时保持模板原有样式
"""

import os
import pandas as pd
from typing import List, Dict, Any, Optional

from core.rowkey_matcher import RowKeyMatcher


class TableFiller:
    """填表执行器。

    根据 matches 中的列映射关系 + 行标识(key)对齐，
    将 src_dfs 的数据填入 tgt_df，支持多源优先级和空模板自动填充。
    """

    # ---- 3遍归一化辅助（与 matcher.py 保持一致，避免循环导入） ----

    @staticmethod
    def _fullwidth_to_halfwidth(text: str) -> str:
        """全角→半角转换。"""
        result = []
        for ch in text:
            code = ord(ch)
            if code == 0x3000:
                result.append(' ')
            elif 0xFF01 <= code <= 0xFF5E:
                result.append(chr(code - 0xFEE0))
            else:
                result.append(ch)
        return ''.join(result)

    @staticmethod
    def _normalize_name(name: str) -> str:
        """标准化列名：全角→半角 + 折叠空白 + 去下划线连字符 + 小写。"""
        if not name:
            return ""
        s = str(name).strip()
        s = TableFiller._fullwidth_to_halfwidth(s)
        s = s.replace('\u3000', ' ')
        s = s.lower()
        # 去除所有空白、下划线、连字符
        import re
        s = re.sub(r'[\s_\-]', '', s)
        return s

    @staticmethod
    def _normalize_name_collapsed(name: str) -> str:
        """标准化列名：全角→半角 + 折叠空白 + 去下划线连字符 + 小写（保留空格折叠）。"""
        if not name:
            return ""
        s = str(name).strip()
        s = TableFiller._fullwidth_to_halfwidth(s)
        import re
        s = re.sub(r'[\s]+', ' ', s)
        s = re.sub(r'[_\-]', '', s)
        s = s.lower()
        return s

    @staticmethod
    def _last_segment(name: str) -> str:
        """取复合字段名的最后一段（> 分隔）。

        例: '一、建筑行业>1.房地产>立项>审批文号' → '审批文号'
            '姓名' → '姓名'
        """
        if not name:
            return ""
        parts = name.rsplit('>', 1)
        return parts[-1].strip()

    @staticmethod
    def _strip_parens(name: str) -> str:
        """去掉字段名中的括号及其内容。

        例: '资金来源（政府投资/企业自筹）' → '资金来源'
            '金额（万元）' → '金额'
        """
        if not name:
            return ""
        import re
        return re.sub(r'[（(][^）)]*[）)]', '', name).strip()

    @staticmethod
    def _coerce_value_for_write(val: Any) -> Any:
        """写回前把「零点时间」的 Timestamp/datetime 转成纯 date。

        修复（V1.3.1 hotfix）：源单元格是纯日期（如 2024/1/1）时，pandas
        读成 '2024-01-01 00:00:00'（Timestamp/datetime），直接写入 openpyxl
        会让 Excel 显示成 '2024/1/1 0:00:00'，与用户数据源格式不一致。

        判定：hour==minute==second==0 即「零点时间」，说明原值本就是日期，
        转成 datetime.date 后 Excel 仅显示日期；非零时刻（如 14:30）保持原样，
        保留时分。None / NaN / 非时间类型原样返回。
        """
        import datetime as _dt
        if val is None:
            return val
        # pd.Timestamp 优先判定（pandas 3.x 下不再是 datetime.datetime 子类，
        # 故必须单独判断，且放在 _dt.datetime 之前以避免漏判）。
        if isinstance(val, pd.Timestamp):
            if pd.isna(val):
                return val
            if val.hour == 0 and val.minute == 0 and val.second == 0:
                return val.date()  # -> datetime.date
            return val
        if isinstance(val, _dt.datetime):
            if val.hour == 0 and val.minute == 0 and val.second == 0:
                return val.date()
            return val
        return val

    # ---- 公开 API ----

    def execute(
        self,
        src_dfs: List[Dict],
        tgt_df: pd.DataFrame,
        matches: List[Dict],
        output_path: Optional[str] = None,
        key_column: Optional[Any] = None,
        src_key_column: Optional[str] = None,
        template_path: str = "",
        header_row: int = 0,
        stop_event: Any = None,
        row_overrides: Optional[Dict] = None,
        progress_cb: Optional[callable] = None,
    ) -> Dict:
        """执行填表（行标识匹配 + 多源优先级 + 空模板自动收集）。

        Args:
            src_dfs: 数据源列表 [{"file":..., "sheet":..., "df": DataFrame}, ...]
                     按添加顺序排列（后添加 = 高优先级）
            tgt_df: 目标模板 DataFrame
            matches: 字段匹配规则列表 [{"tgt_field":..., "src_field":..., "matched":..., ...}, ...]
            output_path: 输出文件路径（可选）
            key_column: 行标识列（列名或索引，None=自动检测）
            template_path: 原始模板文件路径（用于 openpyxl 写回保留格式）

        Returns:
            {"filled": 填充单元格数, "details": [...], "output_path": "..."}
        """
        # ---- 1. 自动检测 key 列 ----
        if key_column is None:
            key_column = self._auto_detect_key_column(tgt_df)
            print(f"[FILLER] 自动检测 key 列: {key_column}")

        # 解析 key 列名
        key_col_name = self._resolve_key_column(tgt_df, key_column)
        print(f"[FILLER] 行标识列: '{key_col_name}'")

        # ---- 2. 提取目标行标识 ----
        tgt_keys = self._extract_keys(tgt_df, key_col_name)
        valid_tgt_keys = [k for k in tgt_keys if k]
        print(f"[FILLER] 目标模板共 {len(tgt_keys)} 行, 有效行标识 {len(valid_tgt_keys)} 个")

        # ---- 2.5 展开纵向合并的 key 列 ----
        # 目标模板中"项目名"等 key 列若为纵向合并单元格（如 A2:A3 同一项目），
        # pandas 读入后仅顶部有值、下方为空，会导致非锚点行被跳过、整行漏填。
        # 复用 ExcelReader 的合并信息，将合并区内下方空值用左上角值回填。
        if key_col_name and template_path and os.path.exists(template_path):
            try:
                tgt_df = self._expand_merged_key_column(
                    tgt_df, template_path, key_col_name, header_row
                )
                tgt_keys = self._extract_keys(tgt_df, key_col_name)
                valid_tgt_keys = [k for k in tgt_keys if k]
                print(
                    f"[FILLER] 合并 key 列展开后: 有效行标识 {len(valid_tgt_keys)} 个"
                )
            except Exception as e:
                print(f"[WARN] 合并 key 列展开失败（已跳过）: {e}")

        # ---- 3. 构建源文件索引（key列映射 + key→行映射） ----
        src_maps = self._build_source_maps(src_dfs, key_col_name, src_key_column)

        # ---- 4. 空模板处理：从源文件收集行标识 ----
        if not valid_tgt_keys:
            print("[FILLER] 目标模板无数据行，启动空模板自动收集...")
            all_keys = self._collect_unique_keys(src_maps)
            if all_keys:
                tgt_df, tgt_keys = self._expand_target_with_keys(
                    tgt_df, key_col_name, all_keys
                )
                valid_tgt_keys = [k for k in tgt_keys if k]
                print(f"[FILLER] 空模板收集完成: 新增 {len(all_keys)} 行")
            else:
                print("[FILLER] 警告: 源文件中也未找到有效行标识，将按顺序填充")

        # ---- 5. 转换列类型，防止 dtype 不兼容写入报错 ----
        # pandas 3.x 下字符串列默认 dtype 名为 'str'（旧版为 'string'），二者及
        # pyarrow 后端均为严格字符串类型，会拒绝 int/float 赋值并抛 TypeError。
        # 统一转 object 后再写入，float/datetime 等均可安全落入。
        for col in tgt_df.columns:
            dt = tgt_df[col].dtype
            dt_name = str(dt)
            if dt_name in ('str', 'string') or (hasattr(dt, 'storage') and dt.storage == 'pyarrow'):
                tgt_df[col] = tgt_df[col].astype(object)
            elif dt_name == 'float64':
                tgt_df[col] = tgt_df[col].astype(object)

        # ---- 5.5 行标识模糊/同义对齐（跨表项目名不同写法也能对应） ----
        # 例如目标"示范家园三期保障性住房" ↔ 源"示范家园三期保障房"。
        # 精确相等自然获得最高分，原行为完整保留；仅当精确未命中时启用模糊。
        rkm = RowKeyMatcher()
        for s_map in src_maps:
            s_map["align"] = rkm.align(tgt_keys, s_map["all_keys"])

        # 应用用户在字段映射页手动做的行级纠偏（覆盖自动配对）
        # row_overrides: {tgt_key: src_key 或 None(解绑)}
        if row_overrides:
            for s_map in src_maps:
                align = s_map.get("align", {})
                for tk, sv in row_overrides.items():
                    if sv is None:
                        align.pop(tk, None)          # 解绑：该目标行不再匹配任何源
                    else:
                        align[tk] = sv                # 强制配对到指定源键
                s_map["align"] = align

        # ---- 6. 执行填充 ----
        # 关键修复：行标识列（key_column / X 轴锚点）绝不能被源字段回填。
        # 否则会把目标自己的行标签（如"示范家园三期保障性住房"）覆盖成源拼写
        # （如"示范家园三期保障房"），破坏行身份。行标识列只用于对齐，不参与填充。
        active_matches = [
            m for m in matches
            if m.get("matched") and m.get("src_field")
            and m.get("tgt_field") != key_col_name
        ]

        filled_cells = 0
        details = []

        for field_idx, m in enumerate(active_matches):
            tgt_field = m["tgt_field"]
            src_field_name = m.get("src_field")

            if tgt_field not in tgt_df.columns:
                details.append({
                    "tgt_field": tgt_field,
                    "status": "skip",
                    "reason": "目标字段不存在于模板中",
                })
                continue

            tgt_col_idx = list(tgt_df.columns).index(tgt_field)
            # 关键修复（V1.2.8）：只从「匹配规则所属源」取列，杜绝跨源列污染。
            # 旧逻辑把 src_field_name 广播到所有源去模糊查找，导致「预算金额」会误把
            # 源1 的「概算金额/总投资」、源3 的近似列当成同名列填进目标，
            # 出现 100 条源记录却填出 244 行的情况。现严格按 m["src_file"]/src_sheet
            # 锁定唯一归属源，其余源该列一律置 None（内层循环遇 None 跳过）。
            target_s_map = self._resolve_source_map_for_match(src_maps, m)
            src_col_for_map = [
                self._find_matching_column(s_map["df"], src_field_name, tgt_field)
                if s_map is target_s_map else None
                for s_map in reversed(src_maps)
            ]
            field_filled = 0
            # V1.2.9 修复：目标模板若含重复行（如同一合同号出现 3 行），同一源行
            # 会被多个 target_key 命中 → 同一源值被复制到多行，造成 100 条源数据
            # 填出 244 格的"通胀"。本字段内已用过的源行不再重复取数。
            used_src_rows: set = set()

            for tgt_row_idx, tgt_key in enumerate(tgt_keys):
                if stop_event is not None and stop_event.is_set():
                    raise TimeoutError("执行超时，已中止填充（未写入文件）")
                if not tgt_key:
                    continue

                # 按源文件逆序遍历（后添加优先）
                found_val = None
                for s_map, pre_src_col in zip(reversed(src_maps), src_col_for_map):
                    if pre_src_col is None:
                        continue
                    # 优先用行对齐结果（含模糊/同义匹配），否则回退精确匹配
                    matched_src = s_map.get("align", {}).get(tgt_key)
                    if matched_src:
                        src_rows = s_map["key_to_row"].get(matched_src)
                    else:
                        src_rows = s_map["key_to_row"].get(tgt_key)
                    if not src_rows:
                        continue

                    ordered_rows = (
                        sorted(
                            src_rows,
                            key=lambda i: -self._count_nonempty(s_map["df"], i),
                        )
                        if len(src_rows) > 1
                        else src_rows
                    )
                    for src_row_idx in ordered_rows:
                        # V1.2.9：源行已被本字段使用过 → 跳过，避免重复填充
                        if (id(s_map["df"]), src_row_idx) in used_src_rows:
                            continue
                        val = s_map["df"].iloc[src_row_idx][pre_src_col]
                        if pd.notna(val) and str(val).strip():
                            found_val = val
                            used_src_rows.add((id(s_map["df"]), src_row_idx))
                            break
                    if found_val is not None:
                        break  # 找到第一个（最高优先级）有数据的源即停止

                if found_val is not None:
                    tgt_df.iloc[tgt_row_idx, tgt_col_idx] = self._coerce_value_for_write(found_val)
                    field_filled += 1

            filled_cells += field_filled
            details.append({
                "tgt_field": tgt_field,
                "src_field": src_field_name,
                "status": "ok",
                "cells": field_filled,
            })

            # 进度回报（驱动前端看门狗与进度条；异常安全）
            if progress_cb:
                try:
                    pct = int((field_idx + 1) / max(1, len(active_matches)) * 100)
                    progress_cb(pct, "填充字段：" + str(tgt_field))
                except Exception:
                    pass

        print(f"[FILLER] 填充完成: {filled_cells} 个单元格, {len(active_matches)} 个匹配规则")

        # ---- 7. 保存结果（原子保存：临时文件 → os.replace，避免 abort 中途留半截损坏文件） ----
        format_converted_note = None
        tmp_path = None
        if output_path:
            try:
                import uuid
                tmp_path = f"{output_path}.xlsx.part_{uuid.uuid4().hex[:8]}"
                tgt_ext = (
                    os.path.splitext(template_path)[1].lower()
                    if (template_path and os.path.exists(template_path))
                    else os.path.splitext(output_path)[1].lower()
                )
                if tgt_ext == ".xls":
                    # openpyxl 不支持写入旧 .xls 格式，自动转存为 .xlsx
                    base, _ = os.path.splitext(output_path)
                    output_path = base + ".xlsx"
                    tgt_df.to_excel(tmp_path, index=False, engine="openpyxl")
                    format_converted_note = "模板为旧 .xls 格式，结果已自动转存为 .xlsx"
                    print(f"[FILLER] {format_converted_note}: {output_path}")
                elif template_path and os.path.exists(template_path):
                    self._save_with_format(tgt_df, template_path, tmp_path, header_row)
                else:
                    tgt_df.to_excel(tmp_path, index=False, engine="openpyxl")
                # 原子改名：abort 事件只可能在 replace 之前触发，故不会留半截文件
                try:
                    os.replace(tmp_path, output_path)
                except (PermissionError, OSError):
                    # 输出文件可能被 Excel / 预览程序占用（共享锁），os.replace 在
                    # Windows 上会报 WinError 5。先尝试删除旧文件再替换一次。
                    if os.path.exists(output_path):
                        try:
                            os.remove(output_path)
                        except OSError:
                            pass
                    os.replace(tmp_path, output_path)
                tmp_path = None
                # 验证文件真的生成
                if not os.path.exists(output_path):
                    raise FileNotFoundError(f"保存后文件不存在: {output_path}")
                file_size = os.path.getsize(output_path)
                print(f"[FILLER] 结果已保存: {output_path} ({file_size} bytes)")
            except Exception as e:
                import traceback
                traceback.print_exc()
                print(f"[FILLER] 保存失败: {e}")
                if tmp_path and os.path.exists(tmp_path):
                    try:
                        os.remove(tmp_path)
                    except Exception:
                        pass
                return {
                    "filled": filled_cells,
                    "details": details,
                    "output_path": output_path or "",
                    "error": f"保存失败: {e}",
                }

        result = {
            "filled": filled_cells,
            "details": details,
            "output_path": output_path or "",
        }
        if format_converted_note:
            result["note"] = format_converted_note
        return result

    # ---- 行级配对计算（供前端可审阅/纠偏列表，不写文件） ----

    def compute_alignment(
        self,
        src_dfs: List[Dict],
        tgt_df: pd.DataFrame,
        src_key: str,
        tgt_key: str,
        template_path: str = "",
        header_row: int = 0,
    ) -> List[Dict]:
        """计算源表与目标表的行级配对，供前端「项目/合同行映射」列表展示。

        Args:
            src_key: 源表用于行标识的列名（如 "项目名称" / "合同编号"）
            tgt_key: 目标表用于行标识的列名（可同源列名不同）
        返回：
            [{"tgt_row_idx", "tgt_key", "src_key", "score", "status"}, ...]
            status: auto(≥阈值 0.82) / suggest(0.6~0.82) / unmatched(<0.6)
        """
        # 目标行标识（展开纵向合并单元格）
        tgt_df2 = tgt_df
        if tgt_key and template_path and os.path.exists(template_path):
            try:
                tgt_df2 = self._expand_merged_key_column(
                    tgt_df, template_path, tgt_key, header_row
                )
            except Exception:
                pass
        if tgt_key and tgt_key in tgt_df2.columns:
            tgt_keys = self._extract_keys(tgt_df2, tgt_key)
        else:
            tgt_keys = [""] * len(tgt_df2)

        # 源行标识（纵向合并列前向填充还原）
        all_src_keys = []
        for info in src_dfs:
            df = info["df"]
            sk = self._find_matching_column(df, src_key) if src_key else None
            df_used = df
            if sk and df[sk].isna().any():
                df_used = df.copy()
                df_used[sk] = df_used[sk].ffill()
            if sk and sk in df_used.columns:
                for idx in range(len(df_used)):
                    val = df_used.iloc[idx][sk]
                    s = str(val).strip() if pd.notna(val) else ""
                    if s and s not in all_src_keys:
                        all_src_keys.append(s)

        rkm = RowKeyMatcher()
        # 先用贪心 1:1 对齐得到高置信度配对（与执行期一致）
        align = rkm.align(tgt_keys, all_src_keys)
        # 预计算归一化源键 + 字符集（V1.3.0 性能优化）：下方未命中兜底循环对每对
        # (tk, sk) 跑 score，规模大时爆炸；复用归一化结果并以「无共享字符则跳过」
        # 的廉价预筛替代，只对可能达阈值(0.85)的候选跑完整 score，正确性不变。
        _norm_src = {}
        _chars_src = {}
        for sk in all_src_keys:
            if not sk or not str(sk).strip():
                continue
            nk = rkm.normalize(sk)
            if nk and nk not in _norm_src:
                _norm_src[nk] = sk
                _chars_src[nk] = set(nk)
        results = []
        for i, tk in enumerate(tgt_keys):
            if not tk:
                continue
            if tk in align:
                results.append({
                    "tgt_row_idx": i,
                    "tgt_key": tk,
                    "src_key": align[tk],
                    "score": round(rkm.score(tk, align[tk]), 2),
                    "status": "auto",
                })
                continue
            # 未命中对齐：给出建议分（0.6~阈值）或标记未匹配
            best_src = None
            best_score = 0.0
            nt = rkm.normalize(tk)
            t_chars = set(nt) if nt else set()
            for nk, sk in _norm_src.items():
                if t_chars and not (t_chars & _chars_src[nk]):
                    continue
                sc = rkm.score(tk, sk)
                if sc > best_score:
                    best_score = sc
                    best_src = sk
            if best_score >= 0.6:
                status = "suggest"
            else:
                status = "unmatched"
            results.append({
                "tgt_row_idx": i,
                "tgt_key": tk,
                "src_key": best_src,
                "score": round(best_score, 2),
                "status": status,
            })
        return results

    # ---- 内部方法 ----

    def _auto_detect_key_column(self, df: pd.DataFrame) -> str:
        """自动检测最佳行标识列。

        扫描所有数据行，返回非空值最多的列的列名。
        若全部为空，返回第一列。

        Args:
            df: 目标 DataFrame

        Returns:
            列名
        """
        if df.empty or len(df.columns) == 0:
            return df.columns[0] if len(df.columns) > 0 else ""

        best_col = df.columns[0]
        best_count = -1

        for col in df.columns:
            # 跳过自动生成的列，但仅作为兜底（避免命名列全空时误选空列）
            is_generated = str(col).startswith("_col_")
            non_empty = 0
            for val in df[col]:
                if pd.notna(val) and str(val).strip():
                    non_empty += 1

            if is_generated:
                if non_empty > best_count and best_count <= 0:
                    best_count = non_empty
                    best_col = col
                continue

            if non_empty > best_count:
                best_count = non_empty
                best_col = col

        return best_col

    def _resolve_key_column(self, df: pd.DataFrame, key_column: Any) -> str:
        """将 key_column（列名或索引）解析为实际的列名。

        Args:
            df: DataFrame
            key_column: 列名(str) 或列索引(int)

        Returns:
            列名
        """
        if isinstance(key_column, int):
            if 0 <= key_column < len(df.columns):
                return df.columns[key_column]
            return df.columns[0]
        if isinstance(key_column, str) and key_column in df.columns:
            return key_column
        # fallback: 尝试匹配
        norm_key = self._normalize_name(key_column)
        for col in df.columns:
            if self._normalize_name(col) == norm_key:
                return col
        return df.columns[0]

    def _extract_keys(self, df: pd.DataFrame, key_col_name: str) -> List[str]:
        """从 DataFrame 中提取行标识列表。

        Args:
            df: DataFrame
            key_col_name: 行标识列名

        Returns:
            行标识字符串列表（按行顺序）
        """
        if key_col_name not in df.columns:
            return [""] * len(df)

        keys = []
        for idx in range(len(df)):
            val = df.iloc[idx][key_col_name]
            if pd.notna(val) and str(val).strip():
                keys.append(str(val).strip())
            else:
                keys.append("")
        return keys

    def _build_source_maps(
        self, src_dfs: List[Dict], key_col_name: str, src_key: Optional[str] = None
    ) -> List[Dict]:
        """为每个源文件构建 key→row 索引映射。

        Args:
            src_dfs: 数据源列表
            key_col_name: 目标表行标识列名
            src_key: 字段映射中对应源字段名（优先使用，解决目标名与源名列名不同问题）

        Returns:
            [{"info":..., "df":..., "key_to_row": {key: row_idx}, "key_col": col_name}, ...]
        """
        maps = []
        for info in src_dfs:
            df = info["df"]
            # 优先使用字段映射中指定的源字段名定位源 key 列；找不到再用目标列名兜底
            src_key_col = None
            if src_key:
                src_key_col = self._find_matching_column(df, src_key)
            if not src_key_col:
                src_key_col = self._find_matching_column(df, key_col_name)
            # 各源独立的兜底解析：多源异构 key 列名时，不再被单一全局 key 名强制
            if not src_key_col:
                last = self._last_segment(str(key_col_name)) if '>' in str(key_col_name) else None
                if last:
                    src_key_col = self._find_matching_column(df, last)
            if not src_key_col:
                _kw = ["项目", "编号", "名称", "合同", "立项", "标识"]
                _best, _bn = None, -1
                for col in df.columns:
                    cs = str(col)
                    if any(k in cs for k in _kw):
                        n = sum(1 for v in df[col] if pd.notna(v) and str(v).strip())
                        if n > _bn:
                            _best, _bn = col, n
                src_key_col = _best
            if not src_key_col:
                src_key_col = self._auto_detect_key_column(df)
            # 纵向合并的 key 列在 pandas 读入后仅顶部有值；前向填充还原合并语义
            if src_key_col:
                col_series = df[src_key_col]
                if col_series.isna().any():
                    df[src_key_col] = col_series.ffill()
            # key_to_row 的值改为「行索引列表」，以支持同一 key 出现多次时保留全部行
            key_to_row = {}
            if src_key_col:
                for idx in range(len(df)):
                    val = df.iloc[idx][src_key_col]
                    key = str(val).strip() if pd.notna(val) else ""
                    if key:
                        if key in key_to_row:
                            key_to_row[key].append(idx)
                            print(
                                f"[WARN] 源 '{info.get('file', '?')}' 中行标识 "
                                f"key='{key}' 重复出现 {len(key_to_row[key])} 次，"
                                f"将按非空单元格更多者优先填充，其余不静默丢弃"
                            )
                        else:
                            key_to_row[key] = [idx]

            maps.append({
                "info": info,
                "df": df,
                "key_col": src_key_col,
                "key_to_row": key_to_row,
                "all_keys": list(key_to_row.keys()),
            })
        return maps

    def _expand_merged_key_column(
        self, df: pd.DataFrame, path: str, col_name: str, header_row: int
    ) -> pd.DataFrame:
        """将目标 key 列的纵向合并单元格展开（下方空值用左上角值回填）。

        解决：目标模板 key 列（如"项目名称"）为 A2:A3 纵向合并时，pandas 仅
        保留顶部值，其余行 key 为空 → 该行在填表时被跳过、整行漏填。
        """
        if col_name not in df.columns:
            return df
        try:
            from core.excel_reader import ExcelReader
        except Exception:
            return df
        reader = ExcelReader()
        try:
            raw_rows, max_row, max_col, _merged = reader._read_raw_with_openpyxl(path)
        except Exception:
            return df
        if header_row is None or header_row < 0 or header_row >= len(raw_rows):
            return df
        hdr = raw_rows[header_row]
        target_norm = self._normalize_name(col_name)
        ac = None
        for c in range(1, max_col + 1):
            v = hdr[c] if c < len(hdr) else None
            if v is None:
                continue
            sv = str(v).strip()
            if self._normalize_name(sv) == target_norm:
                ac = c
                break
            if self._normalize_name(self._last_segment(sv)) == target_norm:
                ac = c
                break
        if ac is None:
            return df
        col_pos = list(df.columns).index(col_name)
        for i in range(len(df)):
            rr = header_row + 1 + i  # 0-based raw 行号（数据行 i）
            if rr >= len(raw_rows):
                break
            row_cells = raw_rows[rr]
            val = row_cells[ac] if ac < len(row_cells) else None
            if val is None or not str(val).strip():
                continue
            cur = df.iloc[i][col_name]
            if pd.isna(cur) or not str(cur).strip():
                df.iloc[i, col_pos] = str(val).strip()
        return df

    def _count_nonempty(self, df: pd.DataFrame, idx: int) -> int:
        """统计某行非空单元格数量（用于重复 key 时优先选信息更完整的一行）。

        性能修复（V1.3.1 hotfix）：原实现在循环内对每个列执行 `df.iloc[idx][col]`，
        等价于每次都重建整行（O(列数²)）。源表多达 120 列且同一 key 重复极多时，
        该函数被疯狂调用，导致每字段 ~20s、整表 ~10 分钟的卡死假象。
        改为一次性取出该行再遍历，复杂度降到 O(列数)，实测每字段 ~0.3s。
        """
        row = df.iloc[idx]
        count = 0
        for v in row:
            if pd.notna(v) and str(v).strip():
                count += 1
        return count

    def _collect_unique_keys(self, src_maps: List[Dict]) -> List[str]:
        """从所有源文件中收集唯一的行标识列表（含模糊去重）。

        当数据源之间存在"同一项目不同写法"（如"示范家园三期保障房"
        与"示范家园三期保障性住房"）时，按相似度聚类合并为同一行标识，
        避免空模板自动收集出重复行。

        Args:
            src_maps: 源文件索引映射列表

        Returns:
            去重（含模糊去重）排序后的行标识列表
        """
        raw = []
        for s_map in src_maps:
            for key in s_map["key_to_row"].keys():
                if key and key.strip():
                    raw.append(key.strip())
        if not raw:
            return []
        rkm = RowKeyMatcher()
        clusters: List[List[str]] = []
        for k in raw:
            placed = False
            for cl in clusters:
                if rkm.score(k, cl[0]) >= rkm.threshold:
                    cl.append(k)
                    placed = True
                    break
            if not placed:
                clusters.append([k])
        return sorted(cl[0] for cl in clusters)

    def _expand_target_with_keys(
        self, tgt_df: pd.DataFrame, key_col_name: str, all_keys: List[str]
    ) -> tuple:
        """将收集到的行标识扩展为目标 DataFrame 的新行。

        Args:
            tgt_df: 原始目标 DataFrame（可能只有标题行）
            key_col_name: 行标识列名
            all_keys: 收集到的行标识列表

        Returns:
            (扩展后的 DataFrame, 新的 tgt_keys 列表)
        """
        import pandas as pd

        # 确保 key 列存在于模板中
        if key_col_name not in tgt_df.columns:
            # key 列不存在，插入为第一列
            tgt_df.insert(0, key_col_name, None)

        new_rows = []
        for key in all_keys:
            row = {col: None for col in tgt_df.columns}
            row[key_col_name] = key
            new_rows.append(row)

        if new_rows:
            new_df = pd.DataFrame(new_rows)
            tgt_df = pd.concat([tgt_df, new_df], ignore_index=True)

        # 重新提取 keys
        tgt_keys = self._extract_keys(tgt_df, key_col_name)
        return tgt_df, tgt_keys

    def _find_matching_column(
        self,
        df: pd.DataFrame,
        primary_name: str,
        fallback_name: str = "",
    ) -> Optional[str]:
        """在 DataFrame 中查找匹配的列名。

        匹配优先级：
          1. 候选名精确匹配
          2. 候选名标准化匹配（3遍归一化）
          3. 候选名 _last_segment 后匹配（候选含 > 时）
          4. 源列 _last_segment 后匹配 candidate（源列含 > 时）
          5. 备选名候选名
          6. 备选名 _last_segment 匹配

        Args:
            df: DataFrame
            primary_name: 首选列名（来自 match 的 src_field）
            fallback_name: 备选列名（来自 match 的 tgt_field）

        Returns:
            匹配到的列名，或 None
        """
        if not primary_name and not fallback_name:
            return None

        candidates = [primary_name]
        if fallback_name and fallback_name != primary_name:
            candidates.append(fallback_name)

        for candidate in candidates:
            if not candidate:
                continue

            # 1. 精确匹配
            if candidate in df.columns:
                return candidate

            # 2. 标准化匹配（同时准备候选 _last_segment）
            norm_cand = self._normalize_name(candidate)
            norm_cand_collapsed = self._normalize_name_collapsed(candidate)
            cand_short = self._last_segment(candidate) if '>' in candidate else None

            for col in df.columns:
                # 2.1 源列标准化匹配
                if self._normalize_name(col) == norm_cand:
                    return col
                if self._normalize_name_collapsed(col) == norm_cand_collapsed:
                    return col

                # 2.2 源列 _last_segment 后匹配 candidate（源列含 > 时）
                if '>' in str(col):
                    col_short = self._last_segment(str(col))
                    if col_short:
                        if col_short == candidate:
                            return col
                        if self._normalize_name(col_short) == norm_cand:
                            return col
                        if self._normalize_name_collapsed(col_short) == norm_cand_collapsed:
                            return col
                        # V1.3.1 新增：末段含括号内容时，去掉括号再试
                        col_stripped = self._strip_parens(col_short)
                        if col_stripped and col_stripped != col_short:
                            if self._normalize_name(col_stripped) == norm_cand:
                                return col
                            # 候选也去括号后对比
                            cand_stripped = self._strip_parens(candidate)
                            if cand_stripped and cand_stripped != candidate:
                                if self._normalize_name(col_stripped) == self._normalize_name(cand_stripped):
                                    return col

                # 2.3 候选 _last_segment 后匹配源列
                if cand_short and self._normalize_name(col) == self._normalize_name(cand_short):
                    return col

            # 3. 回退：候选 _last_segment 精确匹配 df.columns
            if cand_short:
                if cand_short in df.columns:
                    return cand_short

        # 4. 兜底：使用 matcher 做模糊匹配（处理同义词、复合字段名）
        #    例如 key 列 "项目名称>项目名称" 需要对应到源表 "立项信息>立项项目名称>立项项目名称"
        from core.matcher import FieldMatcher
        matcher = FieldMatcher()
        best_score = 0.0
        best_col = None
        for candidate in candidates:
            if not candidate:
                continue
            for col in df.columns:
                if str(col).startswith("_col_"):
                    continue
                score = matcher.score_field(str(col), candidate)
                if score > best_score:
                    best_score = score
                    best_col = col
        if best_col and best_score >= 80:
            return best_col

        return None

    def _resolve_source_map_for_match(self, src_maps, m):
        """根据字段匹配规则的 src_file/src_sheet 定位其唯一归属源。

        V1.2.8 跨源列污染修复：某字段只能从匹配规则指定的那个源取列，
        禁止广播到所有源做模糊查找（否则近似列会被误判为同名列）。

        Returns:
            匹配到的源 map（含 "info"/"df"/"key_to_row" 等），或 None
        """
        import os
        want_file = (m.get("src_file") or "")
        want_sheet = (m.get("src_sheet") or "")

        # 1) 精确匹配 file + sheet（两者均来自同一 info["file"]，通常可直接命中）
        for s_map in src_maps:
            info = s_map.get("info", {}) or {}
            if info.get("file") == want_file and info.get("sheet") == want_sheet:
                return s_map

        # 2) 退化为按文件名(basename)匹配，兼容路径写法差异
        want_base = os.path.basename(want_file)
        if want_base:
            for s_map in src_maps:
                info = s_map.get("info", {}) or {}
                if os.path.basename(str(info.get("file", ""))) == want_base and (
                    not want_sheet or info.get("sheet") == want_sheet
                ):
                    return s_map

        # 3) 都没命中（如规则缺 src_file）：回退到首个能找到该列的源，避免整列漏填，
        #    并打印告警以便发现潜在的归属源缺失。
        print(
            f"[WARN] 字段 '{m.get('tgt_field')}' 的归属源(src_file="
            f"{want_file!r})未精确匹配，回退到首个含该列的源"
        )
        for s_map in reversed(src_maps):
            if self._find_matching_column(
                s_map["df"], m.get("src_field"), m.get("tgt_field")
            ):
                return s_map
        return None

    def _save_with_format(
        self, tgt_df: pd.DataFrame, template_path: str, output_path: str,
        header_row: int = 0
    ):
        """使用 openpyxl 写回数据，保留模板原有样式。

        按列位置对应（df 第0列 → Excel 第1列），不再按名称匹配。

        Args:
            tgt_df: 填充后的 DataFrame
            template_path: 原始模板文件路径
            output_path: 输出文件路径
            header_row: 表头行在 Excel 中的行号（0-based，即 detail_header_row）
        """
        from openpyxl import load_workbook

        wb = load_workbook(template_path)
        ws = wb.active

        # 目标列名（用于精确定位真实字段名行）。保留完整列名，
        # 多级表头(>连接)以末段兜底，避免命中率偏低导致表头行探测失败。
        df_cols = [
            str(c) for c in tgt_df.columns if not str(c).startswith("_col_")
        ]

        def _header_score(row_idx: int) -> int:
            """统计 Excel 第 row_idx 行中与目标列名匹配（完整归一化 或 末段归一化）的
            单元格数量。用于定位真实字段名行，兼容单/多级表头。
            匹配规则与 _col_index_by_name 保持一致（完整名优先，末段兜底）。
            """
            if row_idx < 1 or row_idx > ws.max_row:
                return 0
            hit = 0
            for col_idx in range(1, ws.max_column + 1):
                v = ws.cell(row=row_idx, column=col_idx).value
                if v is None:
                    continue
                sv = str(v).strip()
                if not sv or "Unnamed" in sv:
                    continue
                nv = self._normalize_name(sv)
                if not nv:
                    continue
                # 完整列名匹配
                matched = any(self._normalize_name(c) == nv for c in df_cols)
                if not matched:
                    # 末段匹配（多级表头：父级前缀不参与比对）
                    short = self._normalize_name(self._last_segment(sv))
                    if short:
                        matched = any(
                            self._normalize_name(self._last_segment(c)) == short
                            for c in df_cols
                        )
                if matched:
                    hit += 1
            return hit

        # 真实表头行判定阈值：至少匹配 max(2, 列数*0.5) 个列名。
        need = max(2, int(len(df_cols) * 0.5)) if df_cols else 2

        detail_excel_row = None

        # 【V1.3.0 修复：135 行偏移根因】
        # 旧逻辑直接 `cand = header_row + 1` 并在 `_row_match_count` 通过时
        # 采用，再 `data_start_row = detail_excel_row + 1`。当 header_row（即
        # tgt_header_row，来自 excel_reader 检测 / 用户 manual_zone）被误置为
        # 数据行数 / 表尾行号（如 133/134）时，cand 落在一行「数据行」上，
        # 若该数据行的少量值恰好命中列名预筛阈值，便会被误判为表头，导致
        # data_start_row≈135，前 134 个模板数据行整段空白、数值被写到第 135 行起。
        #
        # 新逻辑：把 header_row 仅作为「候选先验」，并用 tgt_df 实际读出的列名
        # 强校验该行；仅当该行的列名吻合度达到阈值才采用。否则（候选行是数据行、
        # 列名几乎不匹配）→ 进入扫描重探，按「列名匹配最多的那一行」确定真实
        # 字段名行。这样无论 header_row 传入何值，data_start_row 都精确落在
        # 真实首数据行（Excel header_row+2 的语义），前 N 行不再空白。
        # 同时 _header_score 用完整列名比对，修复了多级表头(>连接)因旧
        # _row_match_count 仅用末段名而整表探测失败、数据错误地从第 2 行写起的隐患。
        if header_row is not None and header_row >= 0:
            cand = header_row + 1  # 0-based -> 1-based Excel 行号
            if 1 <= cand <= ws.max_row and _header_score(cand) >= need:
                detail_excel_row = cand

        # 兜底：header_row 候选不通过 / 未提供 → 扫描表头候选区(窗口 200)，
        # 选列名匹配最多的那一行作为字段名行（与 excel_reader 探测同源思路）。
        if detail_excel_row is None:
            detail_excel_row = 1  # 默认兜底
            if df_cols:
                best_row, best_score, best_dense = 1, -1, -1
                for row_idx in range(1, min(ws.max_row + 1, 200)):
                    sc = _header_score(row_idx)
                    # 辅助密度：该行非空单元格数。复合双行表头下，真正
                    # 字段名行(r2 子列名)通常比其上方合并大类行(r1)更"密"
                    # （合并单元格更少 → 非空格更多），也比下方稀疏数据行更
                    # "密"。当多行 _header_score 并列或接近时优先选密度更高者，
                    # 杜绝把"合并大类行 / 数据行"误判为字段名行——这正是
                    # 135 类偏移（数据被写到远离表头的行）的根因之一。
                    dense = 0
                    for c in range(1, ws.max_column + 1):
                        v = ws.cell(row=row_idx, column=c).value
                        if v is not None and str(v).strip() and "Unnamed" not in str(v):
                            dense += 1
                    if (sc > best_score or
                            (sc == best_score and dense > best_dense) or
                            (sc == best_score and dense == best_dense and row_idx > best_row)):
                        best_score, best_dense, best_row = sc, dense, row_idx
                if best_score >= need:
                    detail_excel_row = best_row
                else:
                    # 列名退化（扫描无果）等极端情形：信任调用方的 header_row，
                    # 不再强行扫描，避免把数据写进表头。
                    cand = (header_row + 1) if (
                        header_row is not None and header_row >= 0
                    ) else None
                    detail_excel_row = cand if (
                        cand is not None and 1 <= cand <= ws.max_row
                    ) else 1
            # df_cols 为空（全部为 _col_N）时保持默认 detail_excel_row=1。

        # 数据从表头下一行开始（Excel 1-based）
        data_start_row = detail_excel_row + 1

        # 取消数据区域内的合并单元格，避免 MergedCell 写入报错
        data_end_row = data_start_row + len(tgt_df) - 1
        for mr in list(ws.merged_cells.ranges):
            if mr.max_row >= data_start_row and mr.min_row <= data_end_row:
                ws.unmerge_cells(str(mr))

        # 构建目标表头行（detail_excel_row）的真实列索引，用于按「列名」定位落值单元格。
        # 双行表头 + 合并单元格兼容：前若干列（序号/项目名称等）的字段名位于 R1
        # （与 R2 合并，R2 单元格为 None），其余列字段名位于 R2。故每列取
        # R2 优先、R2 为空时回退 R1（合并单元格左上角值），避免前几列被整列跳过。
        template_headers = []
        r1_row = detail_excel_row - 1 if detail_excel_row >= 2 else detail_excel_row
        for col_idx in range(1, ws.max_column + 1):
            v2 = ws.cell(row=detail_excel_row, column=col_idx).value
            v1 = ws.cell(row=r1_row, column=col_idx).value if r1_row != detail_excel_row else None
            s2 = str(v2).strip() if v2 is not None else ""
            s1 = str(v1).strip() if v1 is not None else ""
            template_headers.append(s2 if s2 else s1)

        def _col_index_by_name(df_col_name: str):
            """按列名在目标表头行中查找真实 1-based Excel 列号。

            匹配顺序：整名列名标准化 → 复合字段末段标准化 →
                      末段去括号内容后标准化（V1.3.1 新增）。
            找不到返回 None（调用方跳过并告警）。
            """
            norm_name = self._normalize_name(df_col_name)
            if norm_name:
                for idx, header in enumerate(template_headers):
                    if self._normalize_name(header) == norm_name:
                        return idx + 1
            short = self._last_segment(df_col_name)
            if short:
                norm_short = self._normalize_name(short)
                if norm_short:
                    for idx, header in enumerate(template_headers):
                        if self._normalize_name(self._last_segment(header)) == norm_short:
                            return idx + 1
                # V1.3.1 新增：末段可能含括号内容（如"资金来源（政府投资/企业自筹）"），
                # 但模板表头只有基名（如"资金来源"）。去掉括号内容后再试一次。
                short_stripped = self._strip_parens(short)
                if short_stripped and short_stripped != short:
                    norm_stripped = self._normalize_name(short_stripped)
                    if norm_stripped:
                        for idx, header in enumerate(template_headers):
                            if self._normalize_name(self._last_segment(header)) == norm_stripped:
                                return idx + 1
            return None

        skipped_cols = []
        for df_col_idx in range(len(tgt_df.columns)):
            col_name = str(tgt_df.columns[df_col_idx])
            if col_name.startswith("_col_"):
                continue
            # 按列名定位真实 Excel 列号，而非依赖 DataFrame 列位置
            excel_col = _col_index_by_name(col_name)
            if excel_col is None:
                skipped_cols.append(col_name)
                continue

            for df_row_idx in range(len(tgt_df)):
                val = tgt_df.iloc[df_row_idx, df_col_idx]
                if pd.notna(val):
                    ws.cell(
                        row=data_start_row + df_row_idx,
                        column=excel_col,
                        value=self._coerce_value_for_write(val),
                    )

        if skipped_cols:
            print(
                "[WARN] 以下列在目标模板表头行未找到对应列，已跳过写入: "
                + ", ".join(skipped_cols)
            )

        wb.save(output_path)
        wb.close()

    def _find_excel_column(
        self, template_headers: List[str], df_col_name: str
    ) -> Optional[int]:
        """在模板列名列表中查找匹配的 Excel 列位置（1-based）。

        Args:
            template_headers: 模板标题行字符串列表
            df_col_name: DataFrame 列名

        Returns:
            Excel 列号（1-based），或 None
        """
        norm_name = self._normalize_name(df_col_name)

        for idx, header in enumerate(template_headers):
            if self._normalize_name(header) == norm_name:
                return idx + 1

        return None
