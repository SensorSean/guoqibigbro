import pandas as pd
import os
from typing import List, Dict, Any, Optional, Tuple

FieldInfo = Dict[str, Any]


class ExcelReader:
    def __init__(self):
        self.src_dfs = []
        self.tgt_df = None
        self.tgt_header_row = 0  # 0-based detail header row of last loaded target

    # ------------------------------------------------------------------
    # Raw cell reading via openpyxl (merge-aware)
    # ------------------------------------------------------------------

    def _read_raw_with_openpyxl(self, path, sheet_name=None):
        """Read all cell values with merged cells expanded.

        Returns:
            rows: List[List[value]] — 0‑based, with [0] as dummy row/col
            max_row: int
            max_col: int
            merged_ranges: List[(min_row, max_row, min_col, max_col, value)]
        """
        from openpyxl import load_workbook
        wb = load_workbook(path, data_only=True, read_only=False)
        if sheet_name:
            ws = wb[sheet_name]
        else:
            ws = wb.active
        max_row = min(ws.max_row, 5000)
        max_col = ws.max_column

        # Collect merge info: (min_row, max_row, min_col, max_col, top-left-value)
        # 回退：若合并区左上角单元格为空，则在合并区内扫描首个非空值，
        # 避免"值写在非左上角"的合并单元格被误判为空。
        merged_ranges: List[Tuple[int, int, int, int, Any]] = []
        for mr in ws.merged_cells.ranges:
            cv = ws.cell(row=mr.min_row, column=mr.min_col).value
            if cv is None or str(cv).strip() == "":
                for rr in range(mr.min_row, mr.max_row + 1):
                    for cc in range(mr.min_col, mr.max_col + 1):
                        v = ws.cell(row=rr, column=cc).value
                        if v is not None and str(v).strip() != "":
                            cv = v
                            break
                    if cv is not None and str(cv).strip() != "":
                        break
            if cv is not None and str(cv).strip() != "":
                merged_ranges.append(
                    (mr.min_row, mr.max_row, mr.min_col, mr.max_col, cv)
                )

        # Build merge_map: (row, col) → expanded value
        merge_map: Dict[Tuple[int, int], Any] = {}
        for r_min, r_max, c_min, c_max, cv in merged_ranges:
            for r in range(r_min, r_max + 1):
                for c in range(c_min, c_max + 1):
                    merge_map[(r, c)] = cv

        rows: List[List[Any]] = []
        for r in range(max_row):
            rd: List[Any] = []
            for c in range(max_col + 1):
                if c == 0:
                    rd.append(None)
                    continue
                ar, ac = r + 1, c
                if (ar, ac) in merge_map:
                    val = merge_map[(ar, ac)]
                else:
                    val = ws.cell(row=ar, column=ac).value
                rd.append(val)
            rows.append(rd)

        wb.close()
        return rows, max_row, max_col, merged_ranges

    def _load_sheet_with_manual_zone(
        self, path: str, sheet_name: str, start_row: int, end_row: int
    ) -> pd.DataFrame:
        """按用户指定的 1‑based 表头区域重新读取单个 Sheet。

        先读原始行（合并单元格已展开），用 [start_row, end_row] 范围内的
        非空单元格构造复合列名，再用 pandas 以 end_row 为 header 读数据，
        最后把列名替换为复合列名。这样即可支持 2~3 行表头 + 局部合并。
        """
        # sheet_name=None 时 pandas.read_excel 会返回 dict（所有 sheet）而非
        # DataFrame，这里解析为实际活动 sheet 名，保持与 _read_raw_with_openpyxl 一致。
        if sheet_name is None:
            try:
                from openpyxl import load_workbook

                _wb = load_workbook(path, read_only=True, data_only=True)
                sheet_name = _wb.active.title
                _wb.close()
            except Exception:
                sheet_name = 0
        raw_rows, max_row, max_col, _ = self._read_raw_with_openpyxl(
            path, sheet_name
        )
        if max_row < 1 or max_col < 1:
            return pd.read_excel(path, sheet_name=sheet_name, header=0)

        start_row = max(1, min(start_row, max_row))
        end_row = max(start_row, min(end_row, max_row))

        def _cell_value(r: int, c: int):
            if r - 1 >= len(raw_rows) or c >= len(raw_rows[r - 1]):
                return None
            v = raw_rows[r - 1][c]
            if v is None:
                return None
            s = str(v).strip()
            if s == "" or "Unnamed" in s:
                return None
            return s

        compound_names = []
        for c in range(1, max_col + 1):
            parts = []
            for r in range(start_row, end_row + 1):
                val = _cell_value(r, c)
                if val and (not parts or parts[-1] != val):
                    parts.append(val)
            if parts:
                compound_names.append(">".join(parts))
            else:
                compound_names.append(f"Col{c}")

        print(
            "[DEBUG] Manual zone compound names (first 5): "
            + str(compound_names[:5])
        )

        df = pd.read_excel(path, sheet_name=sheet_name, header=end_row - 1)
        if len(compound_names) == len(df.columns):
            df.columns = compound_names
        else:
            df = self._clean_columns(df)
        return df

    def _get_header_candidates(
        self, path: str, sheet_name: str = None, max_preview_rows: int = 6
    ) -> List[Dict[str, Any]]:
        """返回可用于手动选择的表头区域候选行（1‑based）。"""
        try:
            raw_rows, max_row, max_col, _ = self._read_raw_with_openpyxl(
                path, sheet_name
            )
        except Exception:
            return []
        if max_row < 1:
            return []

        def _row_preview(r: int) -> str:
            if r - 1 >= len(raw_rows):
                return ""
            cells = [
                str(raw_rows[r - 1][c]).strip()
                for c in range(1, min(max_col, 6) + 1)
                if raw_rows[r - 1][c] is not None
                and str(raw_rows[r - 1][c]).strip() != ""
                and "Unnamed" not in str(raw_rows[r - 1][c])
            ]
            txt = " / ".join(cells[:4])
            return txt[:40] + ("…" if len(txt) > 40 else "")

        limit = min(max_row, max_preview_rows)
        singles = []
        for r in range(1, limit + 1):
            singles.append({
                "value": str(r),
                "label": f"第{r}行" + (f" · {_row_preview(r)}" if _row_preview(r) else ""),
            })

        # 增加常见多行表头范围：1‑2、1‑3、2‑3、2‑4…（支持 2~3 行表头）
        ranges = []
        for i in range(1, limit + 1):
            for j in range(i + 1, min(i + 3, limit + 1)):
                ranges.append({
                    "value": f"{i}-{j}",
                    "label": f"第{i}-{j}行",
                })
        return singles + ranges

    # ------------------------------------------------------------------
    # Multi‑level header zone detection  (NEW — replaces simple row pick)
    # ------------------------------------------------------------------

    def _is_title_row(
        self, row_data: List[Any], threshold: float = 0.5
    ) -> bool:
        """Determine whether a row is a title row.

        A title row is one in which the majority of non-empty cells
        contain the same text — for example a large merged-title
        repeated across every column without actual cell merging.

        Conditions:
          - At least 2 non-empty cells are required (a single cell
            is not a meaningful title row).
          - 关键判别：真实表头行的各字段名**互不相同**（无重复值）；
            而标题/类别行必然有**重复值**（同一文本跨列重复，如
            "基础信息|基础信息|经济指标|经济指标" 或整行重复的报表名）。
            因此仅当存在重复值（某值出现 ≥2 次）且占比 ≥ *threshold*
            时才判为标题行，避免把"2 个不同字段名的真实表头"误判为标题。
        """
        non_empty = [
            str(v).strip()
            for v in row_data
            if v is not None and str(v).strip() and "Unnamed" not in str(v)
        ]
        if len(non_empty) < 2:
            return False
        from collections import Counter
        cnt = Counter(non_empty)
        most_common_count = cnt.most_common(1)[0][1]
        # 无重复值（所有字段名互异）→ 必为真实表头，绝不判为标题
        if most_common_count < 2:
            return False
        return most_common_count / len(non_empty) >= threshold

    def _detect_header_zone(
        self, rows, max_row, max_col, merged_ranges
    ) -> Optional[Dict[str, Any]]:
        """Detect the header *zone* — may span multiple rows of merged categories.

        Returns dict with keys:
            detail_header_row  – 0‑based index of the field‑name row
            category_rows      – list of 0‑based row indices above detail_header_row
            column_names       – list of compound column name strings
            column_count       – effective column count
        Returns None when detection fails (caller falls back to old method).
        """
        if max_row < 1 or max_col < 1:
            return None

        # --- helpers ---------------------------------------------------
        def _is_empty(val: Any) -> bool:
            if val is None:
                return True
            s = str(val).strip()
            return s == "" or "Unnamed" in s

        def _cell_in_merge(r1: int, c1: int) -> bool:
            """Is the 1‑based cell covered by ANY merge?"""
            for r_min, r_max, c_min, c_max, _ in merged_ranges:
                if r_min <= r1 <= r_max and c_min <= c1 <= c_max:
                    return True
            return False

        def _merge_col_span_for_row(r1: int) -> float:
            """Max column‑span ratio (0‑1) of any merge covering this 1‑based row."""
            max_span = 0
            for r_min, r_max, c_min, c_max, _ in merged_ranges:
                if r_min <= r1 <= r_max:
                    span = c_max - c_min + 1
                    if span > max_span:
                        max_span = span
            return max_span / max_col if max_col > 0 else 0.0

        # --- 0. Pre‑scan: identify title rows (all‑same‑value rows) ---
        # These are rows like "某建设集团工程项目基本情况…(单位：万元)"
        # repeated in every column.  They are NOT field names.
        title_scan = min(len(rows), 10)
        title_rows: set = set()
        for r in range(title_scan):
            rd = rows[r] if r < len(rows) else []
            # Build a slice that only includes columns 1..max_col
            row_slice = rd[1 : max_col + 1] if len(rd) > 1 else []
            if self._is_title_row(row_slice):
                title_rows.add(r)

        # --- 1. Per‑row statistics ------------------------------------
        scan_limit = min(len(rows), 200)
        eff_counts: List[int] = []       # effective non‑empty cells per row
        real_counts: List[int] = []      # cells NOT from any merge per row
        uniq_counts: List[int] = []      # unique non‑empty values (V1.3.1)

        for r in range(scan_limit):
            rd = rows[r] if r < len(rows) else []
            ne = 0
            rc = 0
            vals_seen = set()
            for c in range(1, max_col + 1):
                val = rd[c] if c < len(rd) else None
                if not _is_empty(val):
                    ne += 1
                    sv = str(val).strip()
                    vals_seen.add(sv)
                    # A cell is "real" if it is NOT covered by any merge
                    if not _cell_in_merge(r + 1, c):
                        rc += 1
            eff_counts.append(ne)
            real_counts.append(rc)
            uniq_counts.append(len(vals_seen))

        # 标题行必须位于数据/表头之上：若某个被标记为标题的行是其下方
        # 再无有效内容的"最后一行"，则它其实是真实表头（典型如空模板：
        # 仅有"标题行 + 表头行"、无数据行的情况）。此时不应把它排除为候选，
        # 否则会落到真实表头之下去找"更满的下一行"而误判。
        title_rows = {
            r for r in title_rows
            if any(eff_counts[i] > 0 for i in range(r + 1, scan_limit))
        }

        # --- 2. Find the *detail_header_row* ----------------------------
        # IMPORTANT: try `eff_counts` FIRST (not real_counts) because
        # for wide tables (e.g. 120 cols) where the actual header row
        # has ALL cells in Row2-Row3 merges, real_counts=0 for the
        # header row but real_counts=75 for the DATA row. The real
        # pass would falsely pick the data row. The eff pass handles
        # this correctly because it counts non-empty cells regardless
        # of merge status.
        detail_header_row: int = -1

        # 阈值：候选表头行至少需有的非空单元格数。
        # 窄表（列数很少）时 max_col//6 < 3，取下限 1，避免每行最多 2 个非空 → 三通道全跳过 → 永远 None；
        # 宽表（列数多）时 max_col//6 会很大，若直接用作阈值会把"非空单元格较少但确为表头"的行误删，
        # 因此用 min(..., 3) 把上限锁回 3（即原始硬编码值），兼顾窄表修复与宽表不回归。
        header_min_nonempty = max(1, min(max_col // 6, 3))
        real_threshold = header_min_nonempty
        min_nonempty = header_min_nonempty

        def _is_stable(candidate: int, trailing_vals: List[int]) -> bool:
            """Check whether the candidate is a stable header row.

            For a true header row, the following data rows may have far
            fewer non-empty cells (e.g., 120-col wide tables with only
            2-3 cells filled per data row).  If the candidate itself fills
            at least half the columns, accept it immediately.

            Otherwise, require following rows to be similar to the
            candidate (multi-row header case).
            """
            if candidate <= 0:
                return False
            # Strong header: if it fills >= 50% of columns, data rows can be sparse
            if candidate >= max_col * 0.5:
                return True
            for tv in trailing_vals:
                lower = min(tv, candidate)
                upper = max(tv, candidate)
                if upper == 0:
                    return False
                if lower / upper < 0.3:
                    return False
            return True

        def _is_wide_category_row(r1: int) -> bool:
            """行 r1（1-based）涉及的合并中，是否有**跨多列**的宽合并？
            
            宽合并（如 A1:H1）→ 大类行（如"立项信息"跨 8 列）
            窄合并（如 A2:A3）→ 字段名行（每个字段独立一列）
            """
            max_col_span = 0
            for r_min, r_max, c_min, c_max, _ in merged_ranges:
                if r_min <= r1 <= r_max:
                    span = c_max - c_min + 1
                    if span > max_col_span:
                        max_col_span = span
            return max_col_span >= 3  # 跨 ≥3 列视为宽合并

        def _looks_like_data_row(rd: List[Any], mcol: int) -> bool:
            """候选行是否看起来像『数据行/汇总行』而非字段名行。

            命中以下任一特征即判为数据行，跳过其作为表头候选：
              - 非空单元格中 ≥60% 可解析为数值（数值型数据行）
              - 首单元格（或任意单元格）命中汇总关键词
                （合计/总计/小计/汇总/累计/均值/平均）
            用于防止"填满的汇总行"被误判为字段名行。
            """
            ne = [
                v for v in rd[1 : mcol + 1]
                if v is not None and str(v).strip() and "Unnamed" not in str(v)
            ]
            if not ne:
                return False

            def _is_num(x: Any) -> bool:
                t = str(x).strip().replace(",", "").replace(" ", "")
                try:
                    float(t)
                    return True
                except ValueError:
                    return False

            if sum(1 for v in ne if _is_num(v)) / len(ne) >= 0.6:
                return True
            # 仅检查首单元格（key/标签列）：真正的汇总/小计/合计行通常首单元格命中
            # 关键词；若检查所有单元格，会把"含合计/累计/平均等字样的字段名"
            # 误判为汇总行，从而错误跳过真实表头（导致深模板数据行被当成表头）。
            if ne and any(
                a in str(ne[0])
                for a in ("合计", "总计", "小计", "求和", "汇总", "累计", "均值", "平均")
            ):
                return True
            return False

        # 【调整顺序】先尝试 eff_counts（适合宽表合并）
        for r in range(scan_limit):
            if r in title_rows:
                continue
            if eff_counts[r] < min_nonempty:
                continue
            if _looks_like_data_row(rows[r], max_col):
                continue
            # 跳过宽合并大类行：它们不是字段名
            if _is_wide_category_row(r + 1):
                continue
            trailing: List[int] = []
            for i in range(r + 1, scan_limit):
                if i in title_rows:
                    continue
                if eff_counts[i] > 0:
                    trailing.append(eff_counts[i])
                    if len(trailing) >= 2:
                        break
            # 跳过「单格标题行」：这类标题（如只有 A1 有值的报表名）比其下方
            # 的真实表头更稀疏；若候选行本身较稀疏(< 60% 列)且紧邻其后的行更"满"，
            # 则它只是标题，真实表头在下一行。覆盖"标题行 + 表头行"的无合并布局
            # （含空模板：仅标题+表头、无数据行的情况）。
            # 用"相对稀疏 + 自身稀疏"双条件，避免误伤密集宽表或窄表中
            # 真实表头恰比某数据行稀疏的边界情况。
            # 【V1.3.1】例外：若候选行的紧上方是宽合并大类行（跨≥3列的merge），
            # 则本行是二级表头的字段名行——字段名行天然比数据行稀疏，不应跳过。
            # 此类表格结构：大类行(宽合并) → 字段名行 → 数据行。
            if (eff_counts[r] < max_col * 0.6 and trailing
                    and eff_counts[r] < max(trailing)
                    and not (r > 0 and _is_wide_category_row(r))):
                continue
            if _is_stable(eff_counts[r], trailing):
                # V1.3.1 修复：3 行合并表头场景下，当前候选行（如大类行 2）
                # 可能提前通过 _is_stable（因为合并单元格展开后非空够多），
                # 但下一行才是真正的字段名行。此时应优先选择更下方、非空
                # 单元格更多（更"密"）的候选行作为 detail_header_row。
                # 前提：下一行不是 title/wide_category/data_row。
                next_r = r + 1
                if next_r < scan_limit:
                    if (next_r not in title_rows
                            and not _is_wide_category_row(next_r + 1)
                            and not _looks_like_data_row(rows[next_r], max_col)
                            and eff_counts[next_r] >= header_min_nonempty):
                        # 下一行也满足表头条件 → 当唯一值更多时推进（合并行展开后 eff 可能相等但 uniq 一定更少）
                        if uniq_counts[next_r] > uniq_counts[r]:
                            r = next_r
                detail_header_row = r
                break

        if detail_header_row < 0:
            # Fallback: try real_counts with same trailing‑row logic
            for r in range(scan_limit):
                if r in title_rows:
                    continue
                if real_counts[r] < real_threshold:
                    continue
                if _looks_like_data_row(rows[r], max_col):
                    continue
                if _is_wide_category_row(r + 1):
                    continue
                trailing: List[int] = []
                for i in range(r + 1, scan_limit):
                    if i in title_rows:
                        continue
                    if real_counts[i] > 0:
                        trailing.append(real_counts[i])
                        if len(trailing) >= 2:
                            break
                if _is_stable(real_counts[r], trailing):
                    # V1.3.1：3 行表头时优先取更细粒度的下一行
                    next_r = r + 1
                    if next_r < scan_limit:
                        if (next_r not in title_rows
                                and not _is_wide_category_row(next_r + 1)
                                and not _looks_like_data_row(rows[next_r], max_col)
                                and real_counts[next_r] >= real_threshold):
                            if real_counts[next_r] > real_counts[r]:
                                r = next_r
                    detail_header_row = r
                    break

        if detail_header_row < 0:
            # Safety net: no title filter, no threshold filter
            for r in range(scan_limit):
                if real_counts[r] < header_min_nonempty:
                    continue
                if _looks_like_data_row(rows[r], max_col):
                    continue
                trailing: List[int] = []
                for i in range(r + 1, scan_limit):
                    if real_counts[i] > 0:
                        trailing.append(real_counts[i])
                        if len(trailing) >= 2:
                            break
                if _is_stable(real_counts[r], trailing):
                    # V1.3.1：3 行表头时优先取更细粒度的下一行
                    next_r = r + 1
                    if next_r < scan_limit:
                        if (not _looks_like_data_row(rows[next_r], max_col)
                                and real_counts[next_r] >= header_min_nonempty):
                            if real_counts[next_r] > real_counts[r]:
                                r = next_r
                    detail_header_row = r
                    break

        if detail_header_row < 0:
            return None  # fall back to old method

        # --- 3. Category rows (rows above detail_header_row) ----------
        # Exclude title rows from categories.
        category_rows: List[int] = []
        for r in range(detail_header_row):
            if r in title_rows:
                continue
            if eff_counts[r] > 0:
                category_rows.append(r)

        # --- 4. Filter out title rows from categories -----------------
        # A title row is characterised by:
        #   (a) its merge spans ≥ 80 % of columns
        #   (b) it is the FIRST category row
        #   (c) the NEXT category row has a merge spanning < 80 %
        # This prevents filtering legitimate full‑width category rows
        # (like in 多级表头 where every category row spans 100 %).
        filtered_cat: List[int] = []
        for i, r in enumerate(category_rows):
            span_r = _merge_col_span_for_row(r + 1)
            if span_r >= 0.8 and i == 0:
                # Check next category row (if any)
                next_span = 1.0
                if i + 1 < len(category_rows):
                    next_span = _merge_col_span_for_row(
                        category_rows[i + 1] + 1
                    )
                if next_span < 0.8:
                    # This is a title row — skip it
                    continue
            filtered_cat.append(r)
        category_rows = filtered_cat

        # --- 5. Build compound column names ---------------------------
        column_names: List[str] = []
        detail_rd = rows[detail_header_row] if detail_header_row < len(rows) else []

        for c in range(1, max_col + 1):
            prefix_parts: List[str] = []

            for cr in category_rows:
                crd = rows[cr] if cr < len(rows) else []
                val = crd[c] if c < len(crd) else None
                if not _is_empty(val):
                    s = str(val).strip()
                    if not prefix_parts or s != prefix_parts[-1]:
                        prefix_parts.append(s)

            detail_val = detail_rd[c] if c < len(detail_rd) else None
            detail_str = (
                str(detail_val).strip()
                if not _is_empty(detail_val)
                else f"Col{c}"
            )

            if prefix_parts:
                compound = ">".join(prefix_parts) + ">" + detail_str
            else:
                compound = detail_str

            column_names.append(compound)

        # --- 6. Log ---------------------------------------------------
        print(
            "[INFO] Header zone: detail=Row"
            + str(detail_header_row + 1)
            + ", categories="
            + str(len(category_rows))
            + " rows, columns="
            + str(max_col)
        )
        print(
            "[DEBUG] Compound names (first 5): "
            + str(column_names[:5])
        )
        print(
            "[INFO] Compound headers (first 10): "
            + str(column_names[:10])
            + (" ..." if len(column_names) > 10 else "")
        )

        return {
            "detail_header_row": detail_header_row,
            "category_rows": category_rows,
            "column_names": column_names,
            "column_count": max_col,
        }

    # ------------------------------------------------------------------
    # Legacy single‑row header detection (kept as fallback)
    # ------------------------------------------------------------------

    def _find_header_row_v2(self, path, sheet_name=None):
        try:
            raw_rows, max_row, max_col, _merged_ranges = (
                self._read_raw_with_openpyxl(path, sheet_name)
            )
        except Exception as e:
            print("[WARN] openpyxl failed (" + str(e) + "), using fallback")
            return self._find_header_row_fallback(path, sheet_name)

        best_row = 0
        best_score = -1
        keywords_list = [
            "dept", "name", "date", "amount", "id", "type",
            "no", "remark", "status", "qty", "unit", "person",
            "部门", "名称", "日期", "金额", "编号", "类型",
            "序号", "备注", "状态", "数量", "单位", "人员",
            "文号", "审批", "批复", "估算", "描述", "立项",
        ]
        for row_idx in range(min(len(raw_rows), 100)):
            rd = raw_rows[row_idx] if row_idx < len(raw_rows) else []
            ne_count = 0
            text_cells = []
            t_chars = 0
            for val in rd:
                if val is None:
                    continue
                s = str(val).strip()
                if s == "" or "Unnamed" in s:
                    continue
                ne_count += 1
                text_cells.append(s)
                t_chars += len(s)
            if ne_count == 0:
                continue
            score = ne_count * 10
            if ne_count >= 5:
                score += 20
            elif ne_count >= 3:
                score += 10
            avg_len = t_chars / ne_count if ne_count > 0 else 0
            if 2 <= avg_len <= 25:
                score += 15
            elif avg_len <= 40:
                score += 8
            kw_hits = sum(
                1 for kw in keywords_list for tc in text_cells if kw in tc
            )
            score += kw_hits * 10
            all_num = True
            for tc in text_cells:
                try:
                    float(tc.replace(",", "").replace(" ", ""))
                except (ValueError, AttributeError):
                    all_num = False
                    break
            if all_num:
                score -= 50
            print(
                "[DEBUG] Row "
                + str(row_idx + 1)
                + ": "
                + str(ne_count)
                + " cols, avg="
                + str(round(avg_len))
                + ", kw="
                + str(kw_hits)
                + ", score="
                + str(score)
            )
            if score > best_score:
                best_score = score
                best_row = row_idx
        print(
            "[INFO] Header selected: Row "
            + str(best_row + 1)
            + " (score="
            + str(best_score)
            + ")"
        )
        return best_row

    def _find_header_row_fallback(self, path, sheet_name=None):
        best_row = 0
        best_score = -1
        for row in range(31):
            try:
                if sheet_name:
                    df = pd.read_excel(path, sheet_name=sheet_name, header=row)
                else:
                    df = pd.read_excel(path, header=row)
                cols = list(df.columns)
                non_empty = [
                    str(c).strip()
                    for c in cols
                    if str(c).strip() != "" and "Unnamed" not in str(c)
                ]
                if len(non_empty) == 0:
                    continue
                score = len(non_empty) * 10
                for c in non_empty:
                    if 1 <= len(c) <= 30:
                        score += 5
                    if len(c) > 0 and (
                        c[0].isalpha() or "\u4e00" <= c[0] <= "\u9fff"
                    ):
                        score += 3
                if score > best_score:
                    best_score = score
                    best_row = row
            except Exception:
                continue
        return best_row

    # ------------------------------------------------------------------
    # Column cleanup
    # ------------------------------------------------------------------

    def _clean_columns(self, df):
        new_cols = []
        seen = {}
        for col in df.columns:
            cs = str(col).strip()
            if "Unnamed" in cs or cs == "":
                new_cols.append(None)
            else:
                if cs in seen:
                    seen[cs] += 1
                    new_cols.append(cs + "_" + str(seen[cs]))
                else:
                    seen[cs] = 1
                    new_cols.append(cs)
        counter = 1
        final_cols = []
        for c in new_cols:
            if c is None:
                final_cols.append("_col_" + str(counter))
                counter += 1
            else:
                final_cols.append(c)
        df.columns = final_cols
        return df

    # ------------------------------------------------------------------
    # Public API: load_files / load_target
    # ------------------------------------------------------------------

    def load_files(self, file_paths, manual_zones=None):
        all_dfs = []
        xls_load_errors = []
        header_info = {}
        for path in file_paths:
            ext = os.path.splitext(path)[1].lower()
            basename = os.path.basename(path)
            zone = (manual_zones or {}).get(path)
            try:
                if ext == ".csv":
                    df = pd.read_csv(path)
                    df = self._clean_columns(df)
                    nf = len(
                        [c for c in df.columns if not str(c).startswith("_col_")]
                    )
                    print("[INFO] " + basename + "[csv]: " + str(nf) + " fields")
                    all_dfs.append({"file": basename, "sheet": "csv", "df": df})
                elif ext == ".xls":
                    # 老格式 .xls：openpyxl 无法读取，交由 pandas（xlrd 引擎）处理。
                    # 合并单元格不展开（已知局限），表头用 pandas 回退探测。
                    try:
                        xl = pd.ExcelFile(path)
                    except Exception as e:
                        msg = "无法读取 .xls（可能缺少 xlrd 引擎）: " + str(e)
                        print("[WARN] " + msg)
                        xls_load_errors.append(basename + ": " + msg)
                        continue
                    for sname in xl.sheet_names:
                        print(
                            "\n[INFO] === Scanning "
                            + basename
                            + "["
                            + sname
                            + "] (.xls) ==="
                        )
                        try:
                            if zone:
                                start_r = max(1, zone.get("start", 1))
                                end_r = max(start_r, zone.get("end", start_r))
                                detail_hr = end_r - 1
                                detected_start, detected_end = start_r, end_r
                            else:
                                detail_hr = self._find_header_row_fallback(
                                    path, sname
                                )
                                detected_start = detected_end = detail_hr + 1
                            df = pd.read_excel(
                                path, sheet_name=sname, header=detail_hr
                            )
                            df = self._clean_columns(df)
                            if path not in header_info:
                                header_info[path] = {
                                    "detected": {
                                        "start": detected_start,
                                        "end": detected_end,
                                    },
                                    "candidates": [],  # .xls 暂不支持候选行预览
                                }
                        except Exception as e:
                            msg = ".xls 工作表读取失败: " + str(e)
                            print("[WARN] " + msg)
                            xls_load_errors.append(basename + "[" + sname + "]: " + msg)
                            continue
                        nv = len(
                            [c for c in df.columns if not str(c).startswith("_col_")]
                        )
                        nr = len(df)
                        print(
                            "[INFO] "
                            + basename
                            + "["
                            + sname
                            + "]: header=Row"
                            + str(detail_hr + 1)
                            + ", "
                            + str(nv)
                            + " fields, "
                            + str(nr)
                            + " rows"
                        )
                        all_dfs.append({"file": basename, "sheet": sname, "df": df})
                    continue
                else:
                    xl = pd.ExcelFile(path)
                    for sname in xl.sheet_names:
                            print(
                                "\n[INFO] === Scanning "
                                + basename
                                + "["
                                + sname
                                + "] ==="
                            )
    
                            # --- 用户手动指定表头区域优先 ---
                            compound_names = None
                            detail_hr = 0
                            detected_start = 1
                            detected_end = 1
                            if zone:
                                start_r = max(1, zone.get("start", 1))
                                end_r = max(start_r, zone.get("end", start_r))
                                detail_hr = end_r - 1
                                detected_start = start_r
                                detected_end = end_r
                                df = self._load_sheet_with_manual_zone(
                                    path, sname, start_r, end_r
                                )
                            else:
                                # --- Try new multi‑level detection first ---
                                try:
                                    raw_rows, max_row, max_col, merged_ranges = (
                                        self._read_raw_with_openpyxl(path, sname)
                                    )
                                    auto_zone = self._detect_header_zone(
                                        raw_rows, max_row, max_col, merged_ranges
                                    )
                                    if auto_zone is not None:
                                        detail_hr = auto_zone["detail_header_row"]
                                        compound_names = auto_zone["column_names"]
                                        cats = auto_zone.get("category_rows", [])
                                        detected_end = detail_hr + 1
                                        detected_start = (
                                            (cats[0] + 1)
                                            if cats
                                            else detected_end
                                        )
                                    else:
                                        # Fallback to legacy single‑row detection
                                        print(
                                            "[INFO] _detect_header_zone failed, "
                                            "falling back to _find_header_row_v2"
                                        )
                                        detail_hr = self._find_header_row_v2(
                                            path, sname
                                        )
                                        detected_start = detected_end = detail_hr + 1
                                except Exception as e:
                                    print(
                                        "[WARN] Header zone detection error ("
                                        + str(e)
                                        + "), falling back"
                                    )
                                    detail_hr = self._find_header_row_v2(path, sname)
                                    detected_start = detected_end = detail_hr + 1

                                # Read with detected header row
                                df = pd.read_excel(
                                    path, sheet_name=sname, header=detail_hr
                                )

                            if path not in header_info:
                                header_info[path] = {
                                    "detected": {
                                        "start": detected_start,
                                        "end": detected_end,
                                    },
                                    "candidates": self._get_header_candidates(
                                        path, sname
                                    ),
                                }
    
                            # Apply compound column names (or fallback cleaning)
                            if compound_names is not None and len(compound_names) == len(
                                df.columns
                            ):
                                df.columns = compound_names
                                nv = len(
                                    [
                                        c
                                        for c in df.columns
                                        if not str(c).startswith("_col_")
                                    ]
                                )
                            else:
                                df = self._clean_columns(df)
                                nv = len(
                                    [
                                        c
                                        for c in df.columns
                                        if not str(c).startswith("_col_")
                                    ]
                                )
    
                            nt = len(df.columns)
                            nr = len(df)
                            print(
                                "[INFO] "
                                + basename
                                + "["
                                + sname
                                + "]: header=Row"
                                + str(detail_hr + 1)
                                + ", "
                                + str(nv)
                                + "/"
                                + str(nt)
                                + " cols, "
                                + str(nr)
                                + " rows"
                            )
                            vn = [
                                str(c)
                                for c in df.columns
                                if not str(c).startswith("_col_")
                            ]
                            for i, nm in enumerate(vn[:10]):
                                print("       [" + str(i + 1) + "] " + nm)
                            if len(vn) > 10:
                                print("       ... +" + str(len(vn) - 10) + " more")
                            all_dfs.append(
                                {"file": basename, "sheet": sname, "df": df}
                            )
            except Exception as e:
                import traceback
                traceback.print_exc()
                print("[WARN] Cannot read " + basename + ": " + str(e))
                continue
        # 🔴 .xls 修复：若完全没读进任何数据、且存在 .xls 读取失败，
        # 直接抛出明确错误（而非静默跳过），让 UI 能提示用户安装 xlrd / 转存 xlsx。
        if not all_dfs and xls_load_errors:
            raise Exception(
                "未能读取任何文件。.xls 老格式读取失败（通常需要 xlrd 引擎）：\n"
                + "\n".join(xls_load_errors)
                + "\n建议：将 .xls 另存为 .xlsx 后重试，或安装 xlrd（pip install xlrd）。"
            )
        self.src_dfs = all_dfs
        return {"dfs": all_dfs, "header_info": header_info}

    def load_target(self, file_path, manual_zone=None):
        try:
            print("\n[INFO] === Scanning Target Template ===")
            ext = os.path.splitext(file_path)[1].lower()
            detected_start = 1
            detected_end = 1

            # 老格式 .xls：跳过 openpyxl（不支持），直接用 pandas 回退探测表头
            if ext == ".xls":
                compound_names = None
                if manual_zone:
                    start_r = max(1, manual_zone.get("start", 1))
                    end_r = max(start_r, manual_zone.get("end", start_r))
                    detail_hr = end_r - 1
                    detected_start, detected_end = start_r, end_r
                else:
                    try:
                        detail_hr = self._find_header_row_fallback(file_path)
                    except Exception:
                        detail_hr = 0
                    detected_start = detected_end = detail_hr + 1
                self.tgt_df = pd.read_excel(file_path, header=detail_hr)
            else:
                # 用户手动指定表头区域优先
                compound_names = None
                detail_hr = 0
                if manual_zone:
                    start_r = max(1, manual_zone.get("start", 1))
                    end_r = max(start_r, manual_zone.get("end", start_r))
                    detail_hr = end_r - 1
                    detected_start, detected_end = start_r, end_r
                    self.tgt_df = self._load_sheet_with_manual_zone(
                        file_path, None, start_r, end_r
                    )
                else:
                    # Try new multi‑level detection
                    try:
                        raw_rows, max_row, max_col, merged_ranges = (
                            self._read_raw_with_openpyxl(file_path)
                        )
                        auto_zone = self._detect_header_zone(
                            raw_rows, max_row, max_col, merged_ranges
                        )
                        if auto_zone is not None:
                            detail_hr = auto_zone["detail_header_row"]
                            compound_names = auto_zone["column_names"]
                            cats = auto_zone.get("category_rows", [])
                            detected_end = detail_hr + 1
                            detected_start = (
                                (cats[0] + 1) if cats else detected_end
                            )
                        else:
                            print(
                                "[INFO] _detect_header_zone failed, "
                                "falling back to _find_header_row_v2"
                            )
                            detail_hr = self._find_header_row_v2(file_path)
                            detected_start = detected_end = detail_hr + 1
                    except Exception as e:
                        print(
                            "[WARN] Header zone detection error ("
                            + str(e)
                            + "), falling back"
                        )
                        detail_hr = self._find_header_row_v2(file_path)
                        detected_start = detected_end = detail_hr + 1

                    self.tgt_df = pd.read_excel(file_path, header=detail_hr)

            if compound_names is not None and len(compound_names) == len(
                self.tgt_df.columns
            ):
                self.tgt_df.columns = compound_names
            else:
                self.tgt_df = self._clean_columns(self.tgt_df)

            nv = len(
                [
                    c
                    for c in self.tgt_df.columns
                    if not str(c).startswith("_col_")
                ]
            )
            nr = len(self.tgt_df)
            print(
                "[INFO] Target: header=Row"
                + str(detail_hr + 1)
                + ", "
                + str(nv)
                + " fields, "
                + str(nr)
                + " rows"
            )
            vn = [
                str(c)
                for c in self.tgt_df.columns
                if not str(c).startswith("_col_")
            ]
            for i, nm in enumerate(vn[:10]):
                print("       [" + str(i + 1) + "] " + nm)
            if len(vn) > 10:
                print("       ... +" + str(len(vn) - 10) + " more")
            # 记录检测到的目标表头行，供 filler 写回时精确定位（避免依赖脆弱的重探测）
            self.tgt_header_row = detail_hr
            candidates = []
            if ext != ".xls":
                try:
                    candidates = self._get_header_candidates(file_path)
                except Exception:
                    pass
            return {
                "df": self.tgt_df,
                "header_info": {
                    "detected": {"start": detected_start, "end": detected_end},
                    "candidates": candidates,
                },
            }
        except Exception as e:
            raise Exception("Cannot read target: " + str(e))

    # ------------------------------------------------------------------
    # Field extraction helpers (unchanged)
    # ------------------------------------------------------------------

    def get_all_fields(self, dfs_info):
        fields = []
        for info in dfs_info:
            df = info["df"]
            fl = info["file"]
            sh = info["sheet"]
            for col in df.columns:
                cs = str(col).strip()
                if cs.startswith("_col_"):
                    continue
                sample = []
                try:
                    sample = [
                        str(v)[:40]
                        for v in df[col].dropna().head(3).tolist()
                    ]
                except Exception:
                    pass
                fields.append(
                    {
                        "name": cs,
                        "source_file": fl,
                        "source_sheet": sh,
                        "sample_values": sample,
                    }
                )
        print("[INFO] Extracted " + str(len(fields)) + " source fields")
        for i, f in enumerate(fields[:12]):
            print(
                "       src["
                + str(i + 1)
                + "] '"
                + f["name"]
                + "' from "
                + f["source_file"]
            )
        if len(fields) > 12:
            print("       ... +" + str(len(fields) - 12) + " more")
        return fields

    def get_target_fields(self, df):
        result = [
            str(c)
            for c in df.columns
            if not str(c).startswith("_col_")
        ]
        print("[INFO] Extracted " + str(len(result)) + " target fields")
        return result
